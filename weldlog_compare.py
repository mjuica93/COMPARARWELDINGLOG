#!/usr/bin/env python3
"""
WELDLOG Compare — Comparación genérica de versiones de Welding Log
Ejecutar: streamlit run weldlog_compare.py
"""

import io, os, re, tempfile, warnings
import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from pyxlsb import open_workbook as xlsb_open

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════
_DATE_HINTS = {"date", "fecha", "fit up", "paint", "erection", "release"}

def norm_hdr(v):
    if v is None:
        return None
    s = str(v).replace("\n", " ")
    s = re.sub(r" {2,}", " ", s).strip()
    return s or None

def norm(v):
    if v is None:
        return ""
    if isinstance(v, float) and np.isnan(v):
        return ""
    if isinstance(v, (datetime, date)):
        try:
            return str(v.date() if hasattr(v, "date") else v)
        except Exception:
            return str(v)
    return str(v).strip()

def xlsb_date(v):
    if v is None or (isinstance(v, float) and (np.isnan(v) or v == 0)):
        return None
    try:
        return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(v)).date()
    except Exception:
        return v

def is_date_col(col_name):
    cl = col_name.lower()
    return any(p in cl for p in _DATE_HINTS)

def pick(col, left_v, right_v, date_cols):
    """Devuelve (valor_ganador, estado)."""
    rn, ln = norm(right_v), norm(left_v)
    if rn == "" and ln == "":
        return None,    "vacio"
    if rn == "":
        return left_v,  "solo_izq"
    if ln == "":
        return right_v, "der→escribe"
    if rn == ln:
        return left_v,  "igual"
    if col in date_cols:
        try:
            d_r = right_v if isinstance(right_v, (datetime, date)) else datetime.fromisoformat(rn)
            d_l = left_v  if isinstance(left_v,  (datetime, date)) else datetime.fromisoformat(ln)
            return (right_v, "fecha→der") if d_r >= d_l else (left_v, "fecha→izq")
        except Exception:
            return right_v, "fecha→der"
    return right_v, "der→machaca"

# ═══════════════════════════════════════════════════════════════════════════
# LECTURA DE ARCHIVOS
# ═══════════════════════════════════════════════════════════════════════════
def get_sheets_xlsx(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets

def get_sheets_xlsb(file_bytes):
    with tempfile.NamedTemporaryFile(suffix=".xlsb", delete=False) as tmp:
        tmp.write(file_bytes); path = tmp.name
    try:
        with xlsb_open(path) as wb:
            return wb.sheets
    finally:
        os.unlink(path)

def get_sheets(file_bytes, filename):
    return get_sheets_xlsb(file_bytes) if filename.lower().endswith(".xlsb") \
           else get_sheets_xlsx(file_bytes)

@st.cache_data(show_spinner=False)
def load_df_xlsx(file_bytes, sheet, hdr_row):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet,
                       header=hdr_row, engine="openpyxl")
    df = df.dropna(how="all").reset_index(drop=False)
    df = df.rename(columns={"index": "_xl_idx"})
    df.columns = [norm_hdr(c) or f"_c{i}" for i, c in enumerate(df.columns)]
    df.columns = [c.strip() for c in df.columns]
    return df

@st.cache_data(show_spinner=False)
def load_df_xlsb(file_bytes, sheet, hdr_row):
    with tempfile.NamedTemporaryFile(suffix=".xlsb", delete=False) as tmp:
        tmp.write(file_bytes); path = tmp.name
    try:
        rows = []
        with xlsb_open(path) as wb:
            with wb.get_sheet(sheet) as ws:
                for row in ws.rows():
                    rows.append([c.v for c in row])
    finally:
        os.unlink(path)
    headers = [norm_hdr(rows[hdr_row][i]) or f"_c{i}"
               for i in range(len(rows[hdr_row]))]
    df = pd.DataFrame(rows[hdr_row + 1:], columns=headers)
    df = df.dropna(how="all")
    df.columns = [c.strip() for c in df.columns]
    return df

def load_df(file_bytes, filename, sheet, hdr_row):
    if filename.lower().endswith(".xlsb"):
        return load_df_xlsb(file_bytes, sheet, hdr_row)
    return load_df_xlsx(file_bytes, sheet, hdr_row)

def build_key(df, mode, col_single, col_a, col_b):
    """Añade o devuelve columna __KEY__ según el modo elegido."""
    KEY = "__KEY__"
    if mode == "Columna directa":
        if col_single and col_single in df.columns:
            df = df.copy()
            df[KEY] = df[col_single].astype(str).str.strip()
    else:  # Construir desde dos columnas
        if col_a and col_b and col_a in df.columns and col_b in df.columns:
            df = df.copy()
            def _mk(row):
                a = str(row[col_a]).strip()
                b = row[col_b]
                if pd.isna(b) or str(b).strip() in ("", "nan"):
                    return a
                return a + (str(int(b)) if isinstance(b, float) else str(b).strip())
            df[KEY] = df.apply(_mk, axis=1)
    return df

# ═══════════════════════════════════════════════════════════════════════════
# ANÁLISIS
# ═══════════════════════════════════════════════════════════════════════════
KEY     = "__KEY__"
ISO_COL = "__ISO__"

def add_iso_col(df, iso_col_real):
    df = df.copy()
    if iso_col_real and iso_col_real in df.columns:
        df[ISO_COL] = df[iso_col_real].astype(str).str.strip()
    else:
        df[ISO_COL] = "SIN_ISO"
    return df

@st.cache_data(show_spinner=False)
@st.cache_data(show_spinner=False)
def compute_stats(_df_left, _df_right, _integrate_cols):
    """Returns per-ISO stats: new / mod / solo / total_r / total_l."""
    stats = {}

    # ── pre-build key sets per ISO ──────────────────────────────────────────
    r_key_map = (
        _df_right.groupby(ISO_COL)[KEY].apply(set).to_dict()
        if ISO_COL in _df_right.columns and KEY in _df_right.columns else {}
    )
    l_key_map = (
        _df_left.groupby(ISO_COL)[KEY].apply(set).to_dict()
        if ISO_COL in _df_left.columns and KEY in _df_left.columns else {}
    )

    # ── vectorized mod count: merge on KEY, compare fields ──────────────────
    avail = [c for c in _integrate_cols
             if c in _df_left.columns and c in _df_right.columns]
    mod_counts = {}
    if avail and ISO_COL in _df_left.columns:
        r_sub = _df_right[[ISO_COL, KEY] + avail].copy()
        l_sub = _df_left [[KEY] + avail].copy()
        merged = r_sub.merge(l_sub, on=KEY, suffixes=("_r", "_l"))
        has_diff = pd.Series(False, index=merged.index)
        for c in avail:
            cr, cl = f"{c}_r", f"{c}_l"
            if cr in merged.columns and cl in merged.columns:
                vr = merged[cr].apply(norm)
                vl = merged[cl].apply(norm)
                has_diff |= (vr != "") & (vr != vl)
        merged["_diff"] = has_diff
        mod_counts = merged.groupby(ISO_COL)["_diff"].sum().astype(int).to_dict()

    # ── vectorized solo count: keys in left not present in right ────────────
    solo_counts = {}
    if ISO_COL in _df_left.columns and KEY in _df_left.columns and KEY in _df_right.columns:
        right_keys_all = set(_df_right[KEY].unique())
        l_sub2 = _df_left[[ISO_COL, KEY]].copy()
        l_sub2["_solo"] = ~l_sub2[KEY].isin(right_keys_all)
        solo_counts = l_sub2.groupby(ISO_COL)["_solo"].sum().astype(int).to_dict()

    for iso, r_keys in r_key_map.items():
        iso = str(iso).strip()
        if iso in ("", "nan"):
            continue
        l_keys = l_key_map.get(iso, set())
        stats[iso] = {
            "new":     len(r_keys - l_keys),
            "mod":     int(mod_counts.get(iso, 0)),
            "solo":    int(solo_counts.get(iso, 0)),
            "total_r": len(r_keys),
            "total_l": len(l_keys),
        }
    return stats

@st.cache_data(show_spinner=False)
def get_iso_view(_df_left, _df_right, iso, integrate_cols, date_cols):
    right = _df_right[_df_right[ISO_COL] == iso].copy()
    left  = _df_left [_df_left [ISO_COL] == iso].copy() \
            if ISO_COL in _df_left.columns else pd.DataFrame()

    keys_r = set(right[KEY].unique()) if KEY in right.columns else set()
    keys_l = set(left [KEY].unique()) if KEY in left.columns  else set()
    avail  = [c for c in integrate_cols if c in right.columns]
    rows   = []

    # NUEVAS (solo derecha)
    for key in sorted(keys_r - keys_l):
        r = right[right[KEY] == key].iloc[0]
        row = {"Estado": "NUEVA", KEY: key, "_xl_idx": None, "_diff_cols": avail}
        for c in avail:
            if c in r.index:
                row[c] = r[c]
        rows.append(row)

    # COMUNES
    for key in sorted(keys_r & keys_l):
        r_row  = right[right[KEY] == key].iloc[0]
        l_q    = left [left [KEY] == key]
        l_row  = l_q.iloc[0] if not l_q.empty else None
        xl_idx = int(l_q["_xl_idx"].iloc[0]) \
                 if not l_q.empty and "_xl_idx" in l_q.columns else None

        row = {"Estado": "OK", KEY: key, "_xl_idx": xl_idx, "_diff_cols": []}
        if l_row is not None:
            for c in avail:
                if c in l_row.index:
                    row[c] = l_row[c]
        diff_cols = []
        for c in avail:
            lv = l_row[c] if l_row is not None and c in l_row.index else None
            rv = r_row[c] if c in r_row.index else None
            winner, estado = pick(c, lv, rv, date_cols)
            if "escribe" in estado or "machaca" in estado or "fecha→" in estado:
                diff_cols.append(c)
                row[c] = winner
        row["_diff_cols"] = diff_cols
        if diff_cols:
            row["Estado"] = "MODIF."
        rows.append(row)

    # SOLO IZQUIERDA
    for key in sorted(keys_l - keys_r):
        l_q = left[left[KEY] == key]
        if l_q.empty:
            continue
        l_row  = l_q.iloc[0]
        xl_idx = int(l_q["_xl_idx"].iloc[0]) if "_xl_idx" in l_q.columns else None
        row = {"Estado": "SOLO_IZQ", KEY: key, "_xl_idx": xl_idx, "_diff_cols": []}
        for c in avail:
            if c in l_row.index:
                row[c] = l_row[c]
        rows.append(row)

    return pd.DataFrame(rows) if rows else pd.DataFrame()

# ═══════════════════════════════════════════════════════════════════════════
# GUARDAR
# ═══════════════════════════════════════════════════════════════════════════
def save_to_xlsx(left_bytes, left_sheet, left_hdr, changes_existing, changes_new, all_cols):
    wb = load_workbook(io.BytesIO(left_bytes))
    ws = wb[left_sheet]

    hrow = left_hdr + 1  # 1-indexed en openpyxl
    col_map = {}
    for cell in ws[hrow]:
        v = norm_hdr(cell.value)
        if v and v not in col_map:
            col_map[v] = get_column_letter(cell.column)

    # Actualizar existentes
    for xl_idx, col, value in changes_existing:
        cn = norm_hdr(col)
        if cn not in col_map:
            continue
        ex_row = xl_idx + left_hdr + 2
        ws[f"{col_map[cn]}{ex_row}"].value = value

    # Añadir nuevas dentro de la tabla
    if changes_new:
        table = next(iter(ws.tables.values()), None)
        if table:
            _, _, max_col_tbl, max_row_tbl = range_boundaries(table.ref)
            next_row = max_row_tbl + 1
        else:
            next_row = ws.max_row + 1
            max_col_tbl = None

        for row_dict in changes_new:
            for col, val in row_dict.items():
                if col.startswith("_") or col == KEY or col == ISO_COL:
                    continue
                cn = norm_hdr(col)
                if cn not in col_map or norm(val) == "":
                    continue
                ws[f"{col_map[cn]}{next_row}"].value = val
            next_row += 1

        if table:
            tbl_start   = table.ref.split(":")[0]
            tbl_end_col = get_column_letter(max_col_tbl)
            table.ref   = f"{tbl_start}:{tbl_end_col}{next_row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════
CSS = """
<style>
.block-container { padding-top: .8rem !important; }
[data-testid="stSidebar"] { min-width: 420px; max-width: 420px; }

.pill { display:inline-block; padding:1px 8px; border-radius:10px;
        font-size:11px; font-weight:700; letter-spacing:.3px; }
.pill-nueva    { background:#C6EFCE; color:#276221; }
.pill-modif    { background:#FFEB9C; color:#7D5C00; }
.pill-ok       { background:#EBF3FB; color:#2F75B6; }
.pill-solo_izq { background:#DCE6F1; color:#1F4E79; }

.badge-new { background:#C6EFCE; color:#276221; padding:1px 6px;
             border-radius:10px; font-size:11px; font-weight:600; }

.comp-table { width:100%; border-collapse:collapse; font-size:13px; }
.comp-table th { background:#1F4E79; color:white; padding:5px 10px; text-align:left; }
.comp-table td { padding:4px 10px; border-bottom:1px solid #E8E8E8; }
.comp-table tr.diff td.old  { background:#FCE4D6; }
.comp-table tr.diff td.new  { background:#C6EFCE; font-weight:600; }
.comp-table tr.diff td.win  { background:#FFEB9C; font-weight:700; color:#7D5C00; }
.comp-table tr:hover td     { background:#F5F5F5 !important; }

[data-testid="stSidebar"] button {
    padding:3px 8px !important; font-size:12px !important;
    min-height:0 !important; height:auto !important;
}
.file-ok { color:#276221; font-weight:600; font-size:13px; }
.file-no { color:#9E9E9E; font-size:13px; }
</style>
"""

# ═══════════════════════════════════════════════════════════════════════════
# COMPONENTES
# ═══════════════════════════════════════════════════════════════════════════
def render_estado(estado):
    cls = {"NUEVA":"nueva","MODIF.":"modif","OK":"ok","SOLO_IZQ":"solo_izq"}.get(estado,"ok")
    return f'<span class="pill pill-{cls}">{estado}</span>'

def render_comparison_table(r_row, l_row, avail, date_cols):
    rows_html = []
    for col in avail:
        lv = norm(l_row[col]) if l_row is not None and col in l_row.index else ""
        rv = norm(r_row[col]) if r_row is not None and col in r_row.index else ""
        winner, estado = pick(
            col,
            l_row[col] if l_row is not None and col in l_row.index else None,
            r_row[col] if r_row is not None and col in r_row.index else None,
            date_cols,
        )
        win_v = norm(winner)
        if lv == "" and rv == "":
            continue
        has_diff = "escribe" in estado or "machaca" in estado or "fecha→" in estado
        row_cls  = "diff" if has_diff else ""
        rows_html.append(
            f'<tr class="{row_cls}">'
            f'<td><b>{col}</b></td>'
            f'<td class="old">{lv or "—"}</td>'
            f'<td class="new">{rv or "—"}</td>'
            f'<td class="win">{win_v or "—"}</td>'
            f'<td style="font-size:10px;color:#888">{estado}</td>'
            f'</tr>'
        )
    if not rows_html:
        return "<p>Sin diferencias.</p>"
    return (
        '<table class="comp-table"><thead>'
        '<tr><th>Campo</th><th>Izquierda (base)</th>'
        '<th>Derecha (fuente)</th><th>Propuesta</th><th>Regla</th></tr>'
        '</thead><tbody>' + "".join(rows_html) + "</tbody></table>"
    )

# ═══════════════════════════════════════════════════════════════════════════
# APP
# ═══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="WELDLOG Compare",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)
st.markdown(CSS, unsafe_allow_html=True)

ss = st.session_state

# ── defaults de session ────────────────────────────────────────────────────
for k, v in {
    "iso_idx":       0,
    "done_isos":     set(),
    "changes_exist": [],
    "changes_new":   [],
    "detail_key":    None,
    "cfg_ready":     False,
}.items():
    if k not in ss:
        ss[k] = v

# ═══════════════════════════════════════════════════════════════════════════
# PANEL DE CONFIGURACIÓN (expansible en la parte superior)
# ═══════════════════════════════════════════════════════════════════════════
with st.expander("⚙️ Configuración de archivos", expanded=not ss.cfg_ready):
    col_left, col_sep, col_right = st.columns([5, 1, 5])

    # ── IZQUIERDA ──────────────────────────────────────────────────────────
    with col_left:
        st.markdown("### ◀ Archivo Base (izquierda)")
        st.caption("El que quieres actualizar / destino final")
        left_file = st.file_uploader(
            "Cargar archivo base",
            type=["xlsx", "xlsb"],
            key="left_uploader",
            label_visibility="collapsed",
        )
        if left_file:
            left_bytes = left_file.read()
            left_name  = left_file.name
            left_sheets = get_sheets(left_bytes, left_name)
            lc1, lc2 = st.columns(2)
            with lc1:
                left_sheet = st.selectbox("Hoja", left_sheets, key="left_sheet")
            with lc2:
                left_hdr = st.number_input(
                    "Fila cabecera (0=primera)", min_value=0, max_value=30,
                    value=5, key="left_hdr", step=1,
                )
            st.markdown(f'<p class="file-ok">✔ {left_name}</p>', unsafe_allow_html=True)
        else:
            left_bytes = left_name = left_sheet = left_hdr = None
            st.markdown('<p class="file-no">Sin archivo cargado</p>', unsafe_allow_html=True)

    with col_sep:
        st.markdown("<br><br><br><br><br><div style='text-align:center;font-size:32px'>⇄</div>",
                    unsafe_allow_html=True)

    # ── DERECHA ────────────────────────────────────────────────────────────
    with col_right:
        st.markdown("### Archivo Fuente (derecha) ▶")
        st.caption("La versión nueva / origen de los cambios")
        right_file = st.file_uploader(
            "Cargar archivo fuente",
            type=["xlsx", "xlsb"],
            key="right_uploader",
            label_visibility="collapsed",
        )
        if right_file:
            right_bytes  = right_file.read()
            right_name   = right_file.name
            right_sheets = get_sheets(right_bytes, right_name)
            rc1, rc2 = st.columns(2)
            with rc1:
                right_sheet = st.selectbox("Hoja", right_sheets, key="right_sheet")
            with rc2:
                right_hdr = st.number_input(
                    "Fila cabecera (0=primera)", min_value=0, max_value=30,
                    value=4, key="right_hdr", step=1,
                )
            st.markdown(f'<p class="file-ok">✔ {right_name}</p>', unsafe_allow_html=True)
        else:
            right_bytes = right_name = right_sheet = right_hdr = None
            st.markdown('<p class="file-no">Sin archivo cargado</p>', unsafe_allow_html=True)

    # ── CONFIGURACIÓN DE COLUMNAS (solo si ambos cargados) ─────────────────
    if left_bytes and right_bytes:
        st.divider()
        st.markdown("#### Configuración de columnas")

        with st.spinner("Cargando hojas..."):
            df_left_raw  = load_df(left_bytes,  left_name,  left_sheet,  int(left_hdr))
            df_right_raw = load_df(right_bytes, right_name, right_sheet, int(right_hdr))

        left_cols  = [c for c in df_left_raw.columns  if not c.startswith("_")]
        right_cols = [c for c in df_right_raw.columns if not c.startswith("_")]
        all_cols   = sorted(set(left_cols) | set(right_cols))

        kc1, kc2, kc3 = st.columns([2, 3, 3])

        with kc1:
            iso_col_left  = st.selectbox(
                "Columna ISO (izq.)",
                [""] + left_cols,
                index=(left_cols.index("Isometric") + 1) if "Isometric" in left_cols else 0,
                key="iso_col_left",
            )
            iso_col_right = st.selectbox(
                "Columna ISO (der.)",
                [""] + right_cols,
                index=(right_cols.index("Isometric") + 1) if "Isometric" in right_cols else 0,
                key="iso_col_right",
            )

        with kc2:
            st.markdown("**Clave única — Archivo base (izq.)**")
            key_mode_l = st.radio(
                "Modo clave izq.",
                ["Columna directa", "Construir desde 2 columnas"],
                key="key_mode_l", label_visibility="collapsed",
            )
            if key_mode_l == "Columna directa":
                key_col_l_single = st.selectbox(
                    "Columna clave izq.",
                    [""] + left_cols,
                    index=(left_cols.index("ISO+WELDNO-ORIGINAL") + 1)
                          if "ISO+WELDNO-ORIGINAL" in left_cols else 0,
                    key="key_col_l_single",
                )
                key_col_l_a = key_col_l_b = None
            else:
                lkc1, lkc2 = st.columns(2)
                with lkc1:
                    key_col_l_a = st.selectbox(
                        "Columna A (texto)",
                        [""] + left_cols,
                        index=(left_cols.index("Isometric") + 1) if "Isometric" in left_cols else 0,
                        key="key_col_l_a",
                    )
                with lkc2:
                    key_col_l_b = st.selectbox(
                        "Columna B (número)",
                        [""] + left_cols,
                        index=0, key="key_col_l_b",
                    )
                key_col_l_single = None

        with kc3:
            st.markdown("**Clave única — Archivo fuente (der.)**")
            key_mode_r = st.radio(
                "Modo clave der.",
                ["Columna directa", "Construir desde 2 columnas"],
                key="key_mode_r", label_visibility="collapsed",
            )
            if key_mode_r == "Columna directa":
                key_col_r_single = st.selectbox(
                    "Columna clave der.",
                    [""] + right_cols,
                    index=(right_cols.index("ISO+WELDNO-ORIGINAL") + 1)
                          if "ISO+WELDNO-ORIGINAL" in right_cols else 0,
                    key="key_col_r_single",
                )
                key_col_r_a = key_col_r_b = None
            else:
                rkc1, rkc2 = st.columns(2)
                with rkc1:
                    key_col_r_a = st.selectbox(
                        "Columna A (texto)",
                        [""] + right_cols,
                        index=(right_cols.index("Isometric") + 1) if "Isometric" in right_cols else 0,
                        key="key_col_r_a",
                    )
                with rkc2:
                    key_col_r_b = st.selectbox(
                        "Columna B (número)",
                        [""] + right_cols,
                        index=0, key="key_col_r_b",
                    )
                key_col_r_single = None

        # Columnas a integrar y fechas
        st.divider()
        ic1, ic2 = st.columns(2)
        with ic1:
            common_cols = [c for c in left_cols if c in right_cols
                           and not c.startswith("_") and c not in ("Isometric",)]
            integrate_cols = st.multiselect(
                "Columnas a comparar/integrar",
                options=all_cols,
                default=common_cols[:60],
                key="integrate_cols",
            )
        with ic2:
            auto_dates = [c for c in all_cols if is_date_col(c)]
            date_cols = st.multiselect(
                "Columnas de fecha (regla: prevalece la más reciente)",
                options=all_cols,
                default=auto_dates,
                key="date_cols",
            )

        # Botón Aplicar
        st.divider()
        if st.button("✅ Aplicar configuración y comparar", type="primary"):
            ss.cfg_left_bytes     = left_bytes
            ss.cfg_left_name      = left_name
            ss.cfg_left_sheet     = left_sheet
            ss.cfg_left_hdr       = int(left_hdr)
            ss.cfg_right_bytes    = right_bytes
            ss.cfg_right_name     = right_name
            ss.cfg_right_sheet    = right_sheet
            ss.cfg_right_hdr      = int(right_hdr)
            ss.cfg_iso_left       = iso_col_left
            ss.cfg_iso_right      = iso_col_right
            ss.cfg_key_mode_l     = key_mode_l
            ss.cfg_key_col_l_s    = key_col_l_single
            ss.cfg_key_col_l_a    = key_col_l_a
            ss.cfg_key_col_l_b    = key_col_l_b
            ss.cfg_key_mode_r     = key_mode_r
            ss.cfg_key_col_r_s    = key_col_r_single
            ss.cfg_key_col_r_a    = key_col_r_a
            ss.cfg_key_col_r_b    = key_col_r_b
            ss.cfg_integrate_cols = list(integrate_cols)
            ss.cfg_date_cols      = set(date_cols)
            ss.cfg_ready          = True
            ss.iso_idx            = 0
            ss.done_isos          = set()
            ss.changes_exist      = []
            ss.changes_new        = []
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════
# PANTALLA SI AÚN NO HAY CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════════
if not ss.cfg_ready:
    st.info(
        "Carga los dos archivos en el panel de arriba, ajusta la configuración "
        "y pulsa **Aplicar configuración y comparar**."
    )
    st.stop()

# ═══════════════════════════════════════════════════════════════════════════
# CARGAR Y PREPARAR DATAFRAMES
# ═══════════════════════════════════════════════════════════════════════════
with st.spinner("Cargando archivos..."):
    df_left_raw  = load_df(ss.cfg_left_bytes,  ss.cfg_left_name,
                           ss.cfg_left_sheet,  ss.cfg_left_hdr)
    df_right_raw = load_df(ss.cfg_right_bytes, ss.cfg_right_name,
                           ss.cfg_right_sheet, ss.cfg_right_hdr)

# Construir columna KEY
df_left  = build_key(df_left_raw,  ss.cfg_key_mode_l,
                     ss.cfg_key_col_l_s, ss.cfg_key_col_l_a, ss.cfg_key_col_l_b)
df_right = build_key(df_right_raw, ss.cfg_key_mode_r,
                     ss.cfg_key_col_r_s, ss.cfg_key_col_r_a, ss.cfg_key_col_r_b)

if KEY not in df_left.columns or KEY not in df_right.columns:
    st.error("No se pudo construir la columna clave. Revisa la configuración.")
    st.stop()

# Columna ISO
df_left  = add_iso_col(df_left,  ss.cfg_iso_left)
df_right = add_iso_col(df_right, ss.cfg_iso_right)

integrate_cols = [c for c in ss.cfg_integrate_cols
                  if c in df_right.columns and not c.startswith("_")]
date_cols      = ss.cfg_date_cols

# Filtrar filas sin ISO válido
df_right = df_right[df_right[ISO_COL].notna() &
                    (df_right[ISO_COL] != "") &
                    (df_right[ISO_COL] != "nan")]

stats    = compute_stats(df_left, df_right, tuple(integrate_cols))
iso_list = sorted(stats.keys())

if not iso_list:
    st.error("No se encontraron isométricos en el archivo fuente. "
             "Verifica la columna ISO seleccionada.")
    st.stop()

# ═══════════════════════════════════════════════════════════════════════════
# SIDEBAR: navegación de isométricos
# ═══════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🔍 WELDLOG Compare")
    st.caption(
        f"**Izq.:** {ss.cfg_left_name}  \n"
        f"**Der.:** {ss.cfg_right_name}"
    )
    if st.button("🔄 Nueva comparación", use_container_width=True):
        for k in ["cfg_ready", "cfg_left_bytes", "cfg_left_name", "cfg_left_sheet",
                  "cfg_left_hdr", "cfg_right_bytes", "cfg_right_name", "cfg_right_sheet",
                  "cfg_right_hdr", "cfg_iso_left", "cfg_iso_right", "cfg_key_mode_l",
                  "cfg_key_col_l_s", "cfg_key_col_l_a", "cfg_key_col_l_b", "cfg_key_mode_r",
                  "cfg_key_col_r_s", "cfg_key_col_r_a", "cfg_key_col_r_b",
                  "cfg_integrate_cols", "cfg_date_cols",
                  "iso_idx", "done_isos", "changes_exist", "changes_new", "manual_links"]:
            if k in ss:
                del ss[k]
        st.rerun()
    st.divider()
    st.caption(
        f"{len(iso_list)} ISOs · "
        f"{len(df_left):,} costuras base · "
        f"{len(df_right):,} fuente"
    )

    # Cambios pendientes + botón de descarga siempre visible
    pending = len(ss.changes_exist) + len(ss.changes_new)
    if pending:
        st.warning(f"{pending} cambios pendientes", icon="⚠️")
        result_bytes = save_to_xlsx(
            ss.cfg_left_bytes, ss.cfg_left_sheet, ss.cfg_left_hdr,
            ss.changes_exist, ss.changes_new, integrate_cols,
        )
        ts       = datetime.now().strftime("%d%m%Y_%H%M")
        out_name = ss.cfg_left_name.rsplit(".", 1)[0] + f"_integrado_{ts}.xlsx"
        dl_label = "💾 Descargar con cambios"
    else:
        result_bytes = ss.cfg_left_bytes
        out_name     = ss.cfg_left_name.rsplit(".", 1)[0] + "_exportado.xlsx"
        dl_label     = "💾 Descargar archivo base"

    st.download_button(
        dl_label,
        data=result_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
    if pending:
        if st.button("🗑 Descartar cambios", use_container_width=True):
            ss.changes_exist = []
            ss.changes_new   = []
            st.rerun()

    st.divider()

    search = st.text_input("🔍 Buscar ISO", placeholder="ej: 2121-BU10",
                           label_visibility="collapsed")
    filt   = st.radio("Filtro", ["Todos", "Con nuevas", "Pendientes"],
                      horizontal=True, label_visibility="collapsed")

    visible = iso_list
    if search:
        visible = [i for i in visible if search.upper() in i.upper()]
    if filt == "Con nuevas":
        visible = [i for i in visible if stats.get(i, {}).get("new", 0) > 0]
    elif filt == "Pendientes":
        visible = [i for i in visible if i not in ss.done_isos]

    # Selección múltiple
    n_sel = sum(1 for i in visible[:200] if ss.get(f"chk_{i}", False))
    cap_c, sel_c, des_c = st.columns([2, 1, 1])
    with cap_c:
        st.caption(f"{len(visible)} isométricos" + (f" · **{n_sel} ✔**" if n_sel else ""))
    with sel_c:
        if st.button("☑ Todos", use_container_width=True, key="sel_all"):
            for _i in visible[:200]:
                st.session_state[f"chk_{_i}"] = True
            st.rerun()
    with des_c:
        if st.button("☐ Ninguno", use_container_width=True, key="des_all"):
            for _i in visible[:200]:
                st.session_state[f"chk_{_i}"] = False
            st.rerun()
    st.divider()

    if ss.iso_idx >= len(iso_list):
        ss.iso_idx = 0

    for iso in visible[:200]:
        s      = stats.get(iso, {})
        new_n  = s.get("new",  0)
        mod_n  = s.get("mod",  0)
        solo_n = s.get("solo", 0)
        done   = iso in ss.done_isos
        active = iso_list[ss.iso_idx] == iso

        col_chk, col_b, col_pills = st.columns([1, 5, 3])
        with col_chk:
            st.checkbox("", key=f"chk_{iso}", label_visibility="collapsed")
        with col_b:
            label = f"{'✔ ' if done else ''}{iso}"
            if st.button(label, key=f"b_{iso}", use_container_width=True,
                         type="primary" if active else "secondary"):
                ss.iso_idx    = iso_list.index(iso)
                ss.detail_key = None
                st.rerun()
        with col_pills:
            pills = []
            if new_n:
                pills.append(
                    f'<span style="background:#C6EFCE;color:#1a6b22;'
                    f'padding:1px 5px;border-radius:3px;font-size:11px;'
                    f'font-weight:600">+{new_n}</span>'
                )
            if mod_n:
                pills.append(
                    f'<span style="background:#FFEB9C;color:#7d5a00;'
                    f'padding:1px 5px;border-radius:3px;font-size:11px;'
                    f'font-weight:600">~{mod_n}</span>'
                )
            if solo_n:
                pills.append(
                    f'<span style="background:#FFCCCC;color:#8b0000;'
                    f'padding:1px 5px;border-radius:3px;font-size:11px;'
                    f'font-weight:600">!{solo_n}</span>'
                )
            if pills:
                st.markdown(
                    '<div style="margin-top:6px;line-height:1.8">'
                    + " ".join(pills) + "</div>",
                    unsafe_allow_html=True,
                )

# ═══════════════════════════════════════════════════════════════════════════
# ÁREA PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════
iso = iso_list[ss.iso_idx]

nav1, nav2, nav_mid, nav3, nav4 = st.columns([1, 1, 6, 1, 1])
with nav1:
    if st.button("◀ Ant.", disabled=ss.iso_idx == 0):
        ss.iso_idx -= 1; ss.detail_key = None; st.rerun()
with nav2:
    if st.button("Sig. ▶", disabled=ss.iso_idx >= len(iso_list) - 1):
        ss.iso_idx += 1; ss.detail_key = None; st.rerun()
with nav_mid:
    st.markdown(f"## {iso}")
with nav3:
    st.caption(f"{ss.iso_idx + 1} / {len(iso_list)}")
with nav4:
    if iso in ss.done_isos:
        st.success("Integrado")

df_view = get_iso_view(df_left, df_right, iso, integrate_cols, date_cols)

if df_view.empty:
    st.warning("Sin datos para este isométrico.")
    st.stop()

# Métricas
n_new  = (df_view["Estado"] == "NUEVA").sum()
n_mod  = (df_view["Estado"] == "MODIF.").sum()
n_ok   = (df_view["Estado"] == "OK").sum()
n_solo = (df_view["Estado"] == "SOLO_IZQ").sum()

m1, m2, m3, m4, m5 = st.columns(5)
m1.metric("Total fuente",  n_new + n_mod + n_ok)
m2.metric("Nuevas",        n_new,  delta=f"+{n_new}"  if n_new else None, delta_color="normal")
m3.metric("Modificadas",   n_mod,  delta=f"~{n_mod}"  if n_mod else None, delta_color="off")
m4.metric("Sin cambios",   n_ok)
m5.metric("Solo en base",  n_solo)

st.divider()

# ── TABS ──────────────────────────────────────────────────────────────────
tab_cmp, tab_edit, tab_detail = st.tabs([
    "Vista comparación", "Editor de costuras", "Detalle costura",
])

# ═══════════════════════════════════════════════════════════════════════════
# TAB 1: COMPARACIÓN
# ═══════════════════════════════════════════════════════════════════════════
with tab_cmp:
    flt1, _ = st.columns([4, 1])
    with flt1:
        show_filter = st.radio(
            "Filtro",
            ["Todas", f"Nuevas ({n_new})", f"Modificadas ({n_mod})",
             f"Sin cambios ({n_ok})", f"Solo base ({n_solo})"],
            horizontal=True, label_visibility="collapsed",
        )
    estado_map = {
        "Todas":                     None,
        f"Nuevas ({n_new})":         "NUEVA",
        f"Modificadas ({n_mod})":    "MODIF.",
        f"Sin cambios ({n_ok})":     "OK",
        f"Solo base ({n_solo})":     "SOLO_IZQ",
    }
    filtro_est = estado_map.get(show_filter)
    df_f = df_view if filtro_est is None else df_view[df_view["Estado"] == filtro_est]

    # Columnas a mostrar: Estado + KEY + primeros campos de integrate_cols disponibles
    show_cols = ["Estado", KEY] + [c for c in integrate_cols if c in df_f.columns][:18]
    df_display = df_f[show_cols].copy()
    for c in date_cols:
        if c in df_display.columns:
            df_display[c] = df_display[c].apply(norm)
    # Fix float columns that are actually integers (e.g. Weld Nº shows as 16.000000)
    for c in df_display.select_dtypes(include="float").columns:
        if c in date_cols:
            continue
        try:
            mask = df_display[c].notna()
            if mask.any() and (df_display.loc[mask, c] % 1 == 0).all():
                df_display[c] = df_display[c].apply(
                    lambda x: str(int(x)) if pd.notna(x) else ""
                )
        except (TypeError, ValueError):
            pass

    _row_bg = {"NUEVA": "#C6EFCE", "MODIF.": "#FFEB9C", "SOLO_IZQ": "#DCE6F1", "OK": ""}

    def cell_style(df_s):
        out = pd.DataFrame("", index=df_s.index, columns=df_s.columns)
        for i, row in df_s.iterrows():
            estado = row.get("Estado", "")
            bg = _row_bg.get(estado, "")
            if bg:
                out.loc[i, :] = f"background-color:{bg}"
            # Stronger yellow on the key column for modified rows
            if estado == "MODIF." and KEY in df_s.columns:
                out.at[i, KEY] = "background-color:#FFC000;font-weight:bold"
        return out

    st.dataframe(
        df_display.style.apply(cell_style, axis=None),
        use_container_width=True,
        height=400,
        hide_index=True,
        column_config={"Estado": st.column_config.TextColumn("Estado", width=80)},
    )

    st.divider()
    st.markdown("**Integrar isométrico:**")
    a1, a2, a3, a4 = st.columns(4)

    def integrate_iso(mode, iso_name=None):
        target    = iso_name if iso_name is not None else iso
        right_iso = df_right[df_right[ISO_COL] == target]
        left_iso  = df_left [df_left [ISO_COL] == target] if ISO_COL in df_left.columns \
                    else pd.DataFrame()
        keys_r = set(right_iso[KEY].unique()) if KEY in right_iso.columns else set()
        keys_l = set(left_iso [KEY].unique()) if KEY in left_iso.columns  else set()
        avail  = [c for c in integrate_cols if c in right_iso.columns]
        n_u = n_a = 0

        if mode in ("all", "existing"):
            for key_v in keys_r & keys_l:
                r_r  = right_iso[right_iso[KEY] == key_v].iloc[0]
                l_q  = left_iso [left_iso [KEY] == key_v]
                if l_q.empty:
                    continue
                l_r    = l_q.iloc[0]
                xl_idx = int(l_q["_xl_idx"].iloc[0]) \
                         if "_xl_idx" in l_q.columns else None
                for c in avail:
                    lv = l_r[c] if c in l_r.index else None
                    rv = r_r[c] if c in r_r.index else None
                    winner, estado = pick(c, lv, rv, date_cols)
                    if "escribe" in estado or "machaca" in estado or "fecha→" in estado:
                        if xl_idx is not None:
                            ss.changes_exist.append((xl_idx, c, winner))
                n_u += 1

        if mode in ("all", "new"):
            # Fila de referencia del ISO en el archivo base (para rellenar campos vacíos)
            left_ref = left_iso.iloc[0] if not left_iso.empty else None

            for key_v in keys_r - keys_l:
                r_r   = right_iso[right_iso[KEY] == key_v].iloc[0]
                row_d = {c: r_r[c] for c in avail
                         if c in r_r.index and norm(r_r[c]) != ""}
                row_d[KEY] = key_v
                # Rellenar campos vacíos con datos del ISO en el archivo base
                if left_ref is not None:
                    for c in left_ref.index:
                        if c.startswith("_"):
                            continue
                        if norm(row_d.get(c, "")) == "" and norm(left_ref[c]) != "":
                            row_d[c] = left_ref[c]
                ss.changes_new.append(row_d)
                n_a += 1

        ss.done_isos.add(target)
        return n_u, n_a

    with a1:
        if st.button("Integrar TODO", type="primary", use_container_width=True):
            u, a = integrate_iso("all")
            st.success(f"Listo: {u} actualizadas, {a} nuevas.")
    with a2:
        if st.button("Solo actualizar existentes", use_container_width=True):
            u, _ = integrate_iso("existing")
            st.success(f"{u} costuras actualizadas.")
    with a3:
        if st.button("Solo añadir nuevas", use_container_width=True):
            _, a = integrate_iso("new")
            st.success(f"{a} costuras nuevas.")
    with a4:
        if st.button("Marcar revisado", use_container_width=True):
            ss.done_isos.add(iso)
            st.info("Marcado sin cambios.")

    # ── Integración masiva (solo si hay ISOs seleccionados) ────────────────
    sel_bulk = [i for i in iso_list if ss.get(f"chk_{i}", False)]
    if sel_bulk:
        st.divider()
        with st.expander(f"⚡ Integración masiva — {len(sel_bulk)} isométricos seleccionados",
                         expanded=True):
            st.caption(", ".join(sel_bulk[:20])
                       + (f" … y {len(sel_bulk)-20} más" if len(sel_bulk) > 20 else ""))
            bk1, bk2, bk3, bk4 = st.columns(4)
            with bk1:
                if st.button("⚡ Integrar TODO", type="primary",
                             use_container_width=True, key="bulk_all"):
                    total_u = total_a = 0
                    prog = st.progress(0, text="Integrando…")
                    for i, iso_s in enumerate(sel_bulk):
                        u, a = integrate_iso("all", iso_name=iso_s)
                        total_u += u; total_a += a
                        prog.progress((i + 1) / len(sel_bulk), text=f"{iso_s} ({i+1}/{len(sel_bulk)})")
                    prog.empty()
                    st.success(f"✔ {total_u} actualizadas · {total_a} nuevas — {len(sel_bulk)} ISOs")
            with bk2:
                if st.button("Solo actualizar existentes",
                             use_container_width=True, key="bulk_exist"):
                    total_u = 0
                    prog = st.progress(0, text="Actualizando…")
                    for i, iso_s in enumerate(sel_bulk):
                        u, _ = integrate_iso("existing", iso_name=iso_s)
                        total_u += u
                        prog.progress((i + 1) / len(sel_bulk), text=f"{iso_s}")
                    prog.empty()
                    st.success(f"✔ {total_u} costuras actualizadas")
            with bk3:
                if st.button("Solo añadir nuevas",
                             use_container_width=True, key="bulk_new"):
                    total_a = 0
                    prog = st.progress(0, text="Añadiendo nuevas…")
                    for i, iso_s in enumerate(sel_bulk):
                        _, a = integrate_iso("new", iso_name=iso_s)
                        total_a += a
                        prog.progress((i + 1) / len(sel_bulk), text=f"{iso_s}")
                    prog.empty()
                    st.success(f"✔ {total_a} costuras nuevas añadidas")
            with bk4:
                if st.button("Marcar todos revisados",
                             use_container_width=True, key="bulk_mark"):
                    for iso_s in sel_bulk:
                        ss.done_isos.add(iso_s)
                    st.info(f"{len(sel_bulk)} isométricos marcados.")

# ═══════════════════════════════════════════════════════════════════════════
# TAB 2: EDITOR
# ═══════════════════════════════════════════════════════════════════════════
with tab_edit:
    st.caption("Edita valores directamente. Pulsa **Aplicar** para añadir al historial.")

    id_cols  = [c for c in ["Estado", KEY] if c in df_view.columns]
    edit_cols = [c for c in integrate_cols if c in df_view.columns]
    df_ed_src = df_view[list(dict.fromkeys(id_cols + edit_cols))].copy()
    for c in date_cols:
        if c in df_ed_src.columns:
            df_ed_src[c] = df_ed_src[c].apply(norm)

    col_cfg = {"Estado": st.column_config.TextColumn("Estado", width=80, disabled=True),
               KEY:      st.column_config.TextColumn("Clave",  width=160, disabled=True)}
    edited = st.data_editor(
        df_ed_src, use_container_width=True, height=400,
        hide_index=True, num_rows="fixed",
        disabled=id_cols, column_config=col_cfg,
    )

    if st.button("Aplicar ediciones", type="primary"):
        applied = 0
        for orig_row, edit_row in zip(df_ed_src.itertuples(index=False),
                                      edited.itertuples(index=False)):
            key_attr = KEY.replace("_","__")
            key_val  = getattr(edit_row, key_attr, None) \
                       or norm(getattr(orig_row, key_attr, ""))
            if not key_val:
                continue
            l_match  = df_left[df_left[KEY] == str(key_val).strip()] \
                       if KEY in df_left.columns else pd.DataFrame()
            is_new   = l_match.empty
            for col in edit_cols:
                col_attr = re.sub(r"[^a-zA-Z0-9]", "_", col)
                ov = norm(getattr(orig_row, col_attr, ""))
                ev = norm(getattr(edit_row, col_attr, ""))
                if ov != ev:
                    if is_new:
                        found = any(norm(nr.get(KEY,"")) == norm(key_val)
                                    for nr in ss.changes_new)
                        if not found:
                            ss.changes_new.append({KEY: key_val, col: ev})
                        else:
                            for nr in ss.changes_new:
                                if norm(nr.get(KEY,"")) == norm(key_val):
                                    nr[col] = ev
                    else:
                        xl_idx = int(l_match["_xl_idx"].iloc[0])
                        ss.changes_exist.append((xl_idx, col, ev))
                    applied += 1
        st.success(f"{applied} campos añadidos al historial.")

# ═══════════════════════════════════════════════════════════════════════════
# TAB 3: DETALLE COSTURA
# ═══════════════════════════════════════════════════════════════════════════
with tab_detail:
    keys_in_view = df_view[KEY].dropna().tolist() if KEY in df_view.columns else []
    if not keys_in_view:
        st.info("Sin costuras para mostrar.")
    else:
        sel_key = st.selectbox(
            "Selecciona costura",
            keys_in_view,
            format_func=lambda k: (
                f"{k}  [{df_view.loc[df_view[KEY]==k,'Estado'].values[0]}]"
                if not df_view[df_view[KEY]==k].empty else k
            ),
            key="sel_detail_key",
        )
        if sel_key:
            r_q = df_right[df_right[KEY] == sel_key]
            l_q = df_left [df_left [KEY] == sel_key] if KEY in df_left.columns \
                  else pd.DataFrame()
            r_r = r_q.iloc[0] if not r_q.empty else None
            l_r = l_q.iloc[0] if not l_q.empty else None
            estado_sel = df_view.loc[df_view[KEY] == sel_key, "Estado"].values[0] \
                         if not df_view[df_view[KEY] == sel_key].empty else "?"

            d1, d2 = st.columns([1, 4])
            with d1:
                st.markdown(f"**Estado:** {render_estado(estado_sel)}", unsafe_allow_html=True)
                st.markdown(f"**Clave:** `{sel_key}`")
                st.divider()
                if l_r is not None and r_r is not None:
                    if st.button("Aplicar fuente → esta costura", use_container_width=True):
                        avail  = [c for c in integrate_cols if c in r_r.index]
                        xl_idx = int(l_q["_xl_idx"].iloc[0]) \
                                 if "_xl_idx" in l_q.columns else None
                        for c in avail:
                            lv = l_r[c] if c in l_r.index else None
                            rv = r_r[c] if c in r_r.index else None
                            winner, estado = pick(c, lv, rv, date_cols)
                            if ("escribe" in estado or "machaca" in estado
                                    or "fecha→" in estado) and xl_idx is not None:
                                ss.changes_exist.append((xl_idx, c, winner))
                        st.success("Cambios añadidos.")
                elif r_r is not None and l_r is None:
                    if st.button("Añadir costura nueva", use_container_width=True):
                        avail = [c for c in integrate_cols if c in r_r.index]
                        row_d = {c: r_r[c] for c in avail
                                 if c in r_r.index and norm(r_r[c]) != ""}
                        row_d[KEY] = sel_key
                        ss.changes_new.append(row_d)
                        st.success("Costura añadida.")

            with d2:
                if r_r is not None:
                    avail = [c for c in integrate_cols if c in r_r.index]
                    html  = render_comparison_table(r_r, l_r, avail, date_cols)
                    st.markdown(html, unsafe_allow_html=True)
                else:
                    st.info("Esta costura solo existe en el archivo base.")

# ═══════════════════════════════════════════════════════════════════════════
# FOOTER: historial de cambios
# ═══════════════════════════════════════════════════════════════════════════
if ss.changes_exist or ss.changes_new:
    st.divider()
    with st.expander(
        f"📋 Historial pendiente — {len(ss.changes_exist)} actualizaciones · "
        f"{ss.changes_new and len(ss.changes_new) or 0} nuevas"
    ):
        if ss.changes_exist:
            st.markdown("**Actualizaciones en costuras existentes (últimas 50):**")
            st.dataframe(
                pd.DataFrame(ss.changes_exist[-50:], columns=["xl_idx","Columna","Valor nuevo"]),
                hide_index=True, use_container_width=True,
            )
        if ss.changes_new:
            st.markdown(f"**Costuras nuevas: {len(ss.changes_new)}**")
            st.dataframe(
                pd.DataFrame([{KEY: r.get(KEY,"")} | {c: r[c] for c in integrate_cols[:4] if c in r}
                               for r in ss.changes_new]),
                hide_index=True, use_container_width=True,
            )
